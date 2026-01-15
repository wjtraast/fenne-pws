#!/usr/bin/env python3
"""Extract content from Word and Excel files for PWS analysis"""

import sys
from pathlib import Path
from docx import Document
from openpyxl import load_workbook

def extract_docx(filepath):
    """Extract text from a Word document"""
    try:
        doc = Document(filepath)
        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
        return '\n'.join(text)
    except Exception as e:
        return f"Error reading {filepath}: {str(e)}"

def extract_xlsx(filepath):
    """Extract content from an Excel workbook"""
    try:
        wb = load_workbook(filepath, data_only=True)
        text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text.append(f"\n=== SHEET: {sheet_name} ===\n")
            for row in sheet.iter_rows(values_only=True):
                row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                if row_text.strip():
                    text.append(row_text)
        return '\n'.join(text)
    except Exception as e:
        return f"Error reading {filepath}: {str(e)}"

def main():
    base_path = Path("/Users/wietjenk/Projects/LiekeArchicad/src/fenne")

    files_to_extract = {
        "Design Route Template": base_path / "bijlage" / "Ontwerpen - onderzoeksverslag.docx",
        "Design Route Grading": base_path / "bijlage" / "Beoordeling - PWS - ontwerpen.xlsx",
        "Student Report": base_path / "resultaat" / "PWS-KookBoekZiektesFLPT.docx",
    }

    output_dir = Path("/Users/wietjenk/Projects/LiekeArchicad/analysis-output")
    output_dir.mkdir(exist_ok=True)

    for name, filepath in files_to_extract.items():
        if not filepath.exists():
            print(f"File not found: {filepath}")
            continue

        print(f"Extracting: {name} ({filepath.name})")

        if filepath.suffix == '.docx':
            content = extract_docx(filepath)
        elif filepath.suffix == '.xlsx':
            content = extract_xlsx(filepath)
        else:
            content = "Unknown file type"

        # Save to text file
        output_file = output_dir / f"{name.replace(' ', '_')}.txt"
        output_file.write_text(content, encoding='utf-8')
        print(f"  -> Saved to {output_file.name}")

    print(f"\nAll extractions saved to: {output_dir}")

if __name__ == "__main__":
    main()

